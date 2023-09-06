# FICHIERS EXCEL
# .XLSX
# openpyxl

import openpyxl

# workbook = openpyxl.load_workbook("octobre.xlsx")
# print(workbook.sheetnames)
# # sheet = workbook['Feuil1']
# sheet = workbook.active
# # print(sheet)
# # cell = sheet["D1"]
#
# for row in range(1, 7):
#     cell = sheet.cell(row, 2)
#     print(cell.value)
#
# print(sheet.max_row)
# print(sheet.max_column)

wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)


def ajouter_data_depuis_workbook(d, wb):
    wb_sheet = wb.active
    for row in range(2, wb_sheet.max_row):
        nom_article = wb_sheet.cell(row, 1).value
        total_ventes = wb_sheet.cell(row, 4).value
        if nom_article and total_ventes:
            if d.get(nom_article):
                d[nom_article].append(total_ventes)
            else:
                d[nom_article] = [total_ventes]


donnees = {}

ajouter_data_depuis_workbook(donnees, wb1)
ajouter_data_depuis_workbook(donnees, wb2)
ajouter_data_depuis_workbook(donnees, wb3)

wb_sortie = openpyxl.Workbook()

sheet = wb_sortie.active
sheet["A1"] = "Articles"
sheet["B1"] = "Septembre"
sheet["C1"] = "Octobre"
sheet["D1"] = "Decembre"

row = 2

for i in donnees.items():
    nom_article = i[0]
    ventes = i[1]
    sheet.cell(row, 1).value = nom_article
    for index in range(0, len(ventes)):
        sheet.cell(row, 2 + index).value = ventes[index]
    row += 1

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=2, max_col=sheet.max_column, max_row=2)
chart_serie = openpyxl.chart.Series(chart_ref, "total ventes (en €)")
chart = openpyxl.chart.BarChart
chart.title = "Evolution du prix des pommes"
chart.append(chart_serie)

sheet.add_chart(chart, "F2")
wb_sortie.save("total_ventes_trimestre.xlsx")
